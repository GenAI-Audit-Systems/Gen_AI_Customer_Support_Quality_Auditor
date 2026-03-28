# ══════════════════════════════════════════════════════════════════════
# Excel Analytics Export — openpyxl
# 4-sheet enterprise export: Summary, Score Trends, Compliance Log, Leaderboard
# ══════════════════════════════════════════════════════════════════════
import io
from datetime import datetime

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    XL_AVAILABLE = True
except ImportError:
    XL_AVAILABLE = False

# ── Style helpers ──────────────────────────────────────────────────────
def _score_fill(score: float):
    if not XL_AVAILABLE:
        return None
    if score >= 75:
        return PatternFill("solid", fgColor="34d399")
    if score >= 50:
        return PatternFill("solid", fgColor="fbbf24")
    return PatternFill("solid", fgColor="f87171")


def _header_style(ws, row: int, cols: int):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="6366f1")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def generate_excel_report(records) -> bytes:
    """
    Generate a 4-sheet Excel analytics report.
    `records` should be an iterable of processor.AuditResult instances.
    Returns raw XLSX bytes.
    """
    if not XL_AVAILABLE:
        return b"openpyxl not installed"

    wb = Workbook()

    # ── Sheet 1: Audit Summary ────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Audit Summary"
    headers1 = ["ID", "File", "Date", "Overall Score", "Grade",
                 "Sentiment", "Resolution", "Professionalism", "Empathy", "Compliance",
                 "Violations Count", "Executive Summary"]
    ws1.append(headers1)
    _header_style(ws1, 1, len(headers1))

    def grade(s):
        if s >= 90: return "A+"
        if s >= 80: return "A"
        if s >= 70: return "B+"
        if s >= 60: return "B"
        if s >= 50: return "C"
        return "D"

    all_records = list(records)
    for r in all_records:
        a = r.audit_json or {}
        sc = r.overall_score or 0
        metrics = a.get("metrics", {})
        violations = a.get("violations", [])
        date = r.created_at.strftime("%Y-%m-%d %H:%M") if hasattr(r.created_at, "strftime") else str(r.created_at)
        
        # Normalize metrics keys (case-insensitive)
        m_norm = {k.lower(): v for k, v in metrics.items()}
        row = [
            r.id, r.filename, date, sc, grade(sc),
            m_norm.get("sentiment", {}).get("label", r.sentiment or "N/A"),
            m_norm.get("resolution", {}).get("label", "N/A"),
            m_norm.get("professionalism", {}).get("label", "N/A"),
            m_norm.get("empathy", {}).get("label", "N/A"),
            m_norm.get("compliance", {}).get("label", "N/A"),
            len(violations),
            a.get("executive_summary", a.get("summary", ""))
        ]
        ws1.append(row)
        score_cell = ws1.cell(row=ws1.max_row, column=4)
        score_cell.fill = _score_fill(sc)

    for col in ws1.columns:
        ws1.column_dimensions[get_column_letter(col[0].column)].width = 20

    # ── Sheet 2: Violations Log ───────────────────────────────────────
    ws2 = wb.create_sheet("Violations Log")
    headers2 = ["Audit ID", "File", "Severity", "Category", "Description", "Evidence", "Remediation"]
    ws2.append(headers2)
    _header_style(ws2, 1, len(headers2))
    for r in all_records:
        a = r.audit_json or {}
        for v in a.get("violations", []):
            ws2.append([
                r.id, r.filename, 
                v.get("severity", "N/A"),
                v.get("category", "Policy Violation"),
                v.get("message", "Compliance Breach"), # Fixed undefined variable
                v.get("quote", "N/A"),
                v.get("action", "N/A")
            ])
    for col in ws2.columns:
        ws2.column_dimensions[get_column_letter(col[0].column)].width = 25

    # ── Sheet 3: Agent Leaderboard ────────────────────────────────────
    ws3 = wb.create_sheet("Agent Leaderboard")
    headers3 = ["Rank", "Agent (File)", "Audits", "Avg Score", "Risk Level"]
    ws3.append(headers3)
    _header_style(ws3, 1, len(headers3))

    agents = {}
    for r in all_records:
        key = (r.filename or "Unknown")[:40]
        if key not in agents:
            agents[key] = {"total": 0, "score_sum": 0, "violations": 0}
        agents[key]["total"]    += 1
        agents[key]["score_sum"] += r.overall_score or 0
        agents[key]["violations"] += len((r.audit_json or {}).get("violations", []))

    ranked = sorted(
        [(k, v) for k, v in agents.items()],
        key=lambda x: x[1]["score_sum"] / max(x[1]["total"], 1),
        reverse=True,
    )
    for rank, (name, v) in enumerate(ranked, 1):
        n    = v["total"]
        avg  = round(v["score_sum"] / n, 1)
        risk = "HIGH" if v["violations"]/n > 1 else "MEDIUM" if avg < 65 else "LOW"
        ws3.append([rank, name, n, avg, risk])
        ws3.cell(row=ws3.max_row, column=4).fill = _score_fill(avg)
    for col in ws3.columns:
        ws3.column_dimensions[get_column_letter(col[0].column)].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
