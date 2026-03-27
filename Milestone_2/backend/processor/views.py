import requests
import json
import os
import re
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from django.http import HttpResponse
from .models import AuditResult
from .utils import split_transcript_by_speaker
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from dotenv import load_dotenv
import io

# Load .env file
load_dotenv()

# ==============================
# API KEYS (From .env)
# ==============================
DEEPGRAM_API_KEY = os.getenv("DEEPGRAM_API_KEY")
OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY")

OPENROUTER_URL = "https://openrouter.ai/api/v1/chat/completions"
MODEL_NAME = "openai/gpt-3.5-turbo"

# ==============================
# TRANSCRIBE USING DEEPGRAM (With Diarization)
# ==============================

def transcribe_with_deepgram(file_content):
    url = "https://api.deepgram.com/v1/listen?diarize=true&smart_format=true&model=nova-2"
    
    headers = {
        "Authorization": f"Token {DEEPGRAM_API_KEY}",
        "Content-Type": "audio/*"
    }

    try:
        response = requests.post(url, headers=headers, data=file_content, timeout=60)
    except requests.exceptions.Timeout:
        raise Exception("Deepgram API request timed out. Please check your network.")
    except requests.exceptions.ConnectionError:
        raise Exception("Could not connect to Deepgram API. Please check your internet connection.")

    if response.status_code != 200:
        raise Exception(f"Deepgram Error: {response.text}")

    result = response.json()
    
    # Process Diarization
    try:
        paragraphs = result.get("results", {}).get("channels", [{}])[0].get("alternatives", [{}])[0].get("paragraphs", {}).get("paragraphs", [])
        if paragraphs:
            dialogue = []
            full_text = ""
            for para in paragraphs:
                speaker_id = para.get('speaker', 0)
                speaker_label = "Agent" if speaker_id == 0 else "Customer"
                sentences = para.get("sentences", [])
                para_text = " ".join([s.get("text", "") for s in sentences])
                dialogue.append({
                    "speaker": speaker_label,
                    "text": para_text
                })
                full_text += f"[{speaker_label}]: {para_text}\n"
            return {"full_text": full_text, "dialogue": dialogue}
    except Exception as e:
        print(f"Diarization fallback: {e}")

    transcript = result["results"]["channels"][0]["alternatives"][0]["transcript"]
    return {"full_text": transcript, "dialogue": [{"speaker": "Unknown", "text": transcript}]}


# ==============================
# QUALITY AUDIT ANALYSIS (Detailed Scoring)
# ==============================

def perform_quality_audit(input_text):
    headers = {
        "Authorization": f"Bearer {OPENROUTER_API_KEY}",
        "Content-Type": "application/json",
        "HTTP-Referer": "http://localhost",
        "X-Title": "AI Quality Auditor"
    }

    system_prompt = """
You are a senior Quality Auditor and Compliance Officer for a customer support center.
Analyze the provided conversation VERY CAREFULLY for quality AND compliance violations.

COMPLIANCE RED FLAGS you MUST detect (list ALL that apply in compliance_issues):
- PII/Data exposure: Agent asks for or accepts full credit card numbers, SSNs, passwords
- Policy misrepresentation: Agent contradicts official policies (e.g. wrong refund timelines)
- Coercion/intimidation: Agent discourages complaints, threatens delays, pressures customer
- Lack of verification: Agent skips identity verification before accessing accounts
- Unprofessional conduct: Dismissive language, belittling the customer's concerns
- Unauthorized promises: Agent offers deals/workarounds outside standard procedure
- Missing confirmation: Agent fails to provide confirmation emails/reference numbers
- Escalation refusal: Agent refuses to transfer to supervisor when requested

Score STRICTLY — a conversation with multiple violations should get compliance 1-3/10, overall_score below 30.

STRICT JSON STRUCTURE (all fields required):
{
  "summary": "2-3 sentence summary of the interaction",
  "scores": {
    "empathy": 0-10,
    "resolution": 0-10,
    "professionalism": 0-10,
    "compliance": 0-10
  },
  "metric_justifications": {
    "empathy": "1-2 sentence explanation with evidence",
    "resolution": "1-2 sentence explanation with evidence",
    "professionalism": "1-2 sentence explanation with evidence",
    "compliance": "1-2 sentence explanation with evidence"
  },
  "overall_score": 0-100,
  "sentiment": "Positive/Neutral/Negative",
  "agent_performance": "Excellent/Good/Satisfactory/Needs Improvement/Poor",
  "call_outcome": "Resolved/Partially Resolved/Unresolved/Escalated",
  "key_findings": ["Finding 1", "Finding 2", "Finding 3"],
  "improvement_tips": ["Tip 1", "Tip 2", "Tip 3"],
  "compliance_issues": ["List EVERY SINGLE compliance violation detected. THIS MUST NOT BE EMPTY IF compliance SCORE < 10"]
}
Only return valid JSON. No extra text.
"""

    payload = {
        "model": MODEL_NAME,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": input_text}
        ],
        "temperature": 0.3,
        "response_format": { "type": "json_object" }
    }

    try:
        response = requests.post(OPENROUTER_URL, headers=headers, data=json.dumps(payload), timeout=60)
    except requests.exceptions.Timeout:
        raise Exception("OpenRouter API request timed out.")
    except requests.exceptions.ConnectionError:
        raise Exception("Could not connect to OpenRouter API.")

    if response.status_code != 200:
        raise Exception(f"OpenRouter Error: {response.text}")

    result = response.json()
    content = result["choices"][0]["message"]["content"]
    return json.loads(content)

# ==============================
# API VIEWS
# ==============================

class AudioProcessView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, *args, **kwargs):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({"error": "No file uploaded"}, status=400)
        
        try:
            # 1. Transcription
            transcription_data = transcribe_with_deepgram(file_obj.read())
            
            # 2. RAG Audit
            from rag.rag_engine import get_engine
            engine = get_engine()
            audit_results = engine.perform_rag_audit(transcription_data["full_text"], "default")
            
            # 3. Persistence
            audit_record = AuditResult.objects.create(
                source_type='audio',
                filename=file_obj.name,
                transcript_json=transcription_data,
                audit_json=audit_results,
                overall_score=audit_results.get("overall_score", 0),
                sentiment=audit_results.get("sentiment", "Neutral")
            )
            
            # TRIGGER ALERTS ENGINE automatically
            try:
                from alerts.alert_engine import evaluate_audit
                evaluate_audit(audit_record.id, audit_results, file_obj.name)
            except Exception as e:
                print(f"[AlertEngine] Failed to generate alerts: {e}")
            
            return Response({
                "id": audit_record.id,
                "source": "audio",
                "transcript": transcription_data,
                "audit": audit_results,
                "policy_context": audit_results.get("policy_context", []),
                "rag_coverage": audit_results.get("rag_coverage", 0.0)
            })
        except Exception as e:
            return Response({"error": str(e)}, status=500)

class TextProcessView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, *args, **kwargs):
        content = request.data.get('content')
        file_obj = request.FILES.get('file')

        if file_obj:
            try:
                content = file_obj.read().decode('utf-8')
            except Exception:
                return Response({"error": "Could not read text file"}, status=400)
        
        if not content:
            return Response({"error": "No content provided"}, status=400)
        
        try:
            # Audit
            audit_results = perform_quality_audit(content)
            
            # Speaker Splitting (if applicable)
            transcript_data = split_transcript_by_speaker(content)
            
            # Persistence
            audit_record = AuditResult.objects.create(
                source_type='text',
                filename=file_obj.name if file_obj else "Direct Input",
                transcript_json=transcript_data,
                audit_json=audit_results,
                overall_score=audit_results.get("overall_score", 0),
                sentiment=audit_results.get("sentiment", "Neutral")
            )
            
            # TRIGGER ALERTS ENGINE automatically
            try:
                from alerts.alert_engine import evaluate_audit
                evaluate_audit(audit_record.id, audit_results, file_obj.name if file_obj else "Direct Input")
            except Exception as e:
                print(f"[AlertEngine] Failed to generate alerts: {e}")
                
            return Response({
                "id": audit_record.id,
                "source": "text",
                "transcript": transcript_data,
                "audit": audit_results
            })
        except Exception as e:
            return Response({"error": str(e)}, status=500)

class HistoryListView(APIView):
    def get(self, request):
        history = AuditResult.objects.all().order_by('-created_at')[:10]
        data = [{
            "id": h.id,
            "source": h.source_type,
            "filename": h.filename,
            "score": h.overall_score,
            "sentiment": h.sentiment,
            "date": h.created_at.strftime('%Y-%m-%d %H:%M'),
            "audit": h.audit_json,
            "transcript": h.transcript_json
        } for h in history]
        return Response(data)

class DiagnosticView(APIView):
    def get(self, request):
        results = {}
        # Test Deepgram
        try:
            dg_resp = requests.get("https://api.deepgram.com/v1/projects", 
                                  headers={"Authorization": f"Token {DEEPGRAM_API_KEY}"}, 
                                  timeout=10)
            results["deepgram"] = "Connected" if dg_resp.status_code in [200, 401, 403] else f"Error: {dg_resp.status_code}"
        except Exception as e:
            results["deepgram"] = f"Timeout/Error: {str(e)}"

        # Test OpenRouter
        try:
            or_resp = requests.get("https://openrouter.ai/api/v1/models", timeout=10)
            results["openrouter"] = "Connected" if or_resp.status_code == 200 else f"Error: {or_resp.status_code}"
        except Exception as e:
            results["openrouter"] = f"Timeout/Error: {str(e)}"
            
        # Test Neon
        db_url = os.getenv("DATABASE_URL")
        results["neon"] = "Connected" if db_url and "neon.tech" in db_url else "Disconnected (Fallback to SQLite)"

        # Test Milvus
        from rag.milvus_client import MILVUS_AVAILABLE
        results["milvus"] = "Connected" if MILVUS_AVAILABLE else "Disconnected (Mock Mode)"

        return Response(results)

class ExcelExportView(APIView):
    """
    GET /api/export-excel/<int:audit_id>/ - Generates a professional Excel audit report.
    """
    def get(self, request, audit_id):
        try:
            record = AuditResult.objects.get(id=audit_id)
            audit = record.audit_json
            transcript = record.transcript_json
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Audit Report"
            
            # Header Styling
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
            align_center = Alignment(horizontal="center", vertical="center")
            
            # 1. Summary Section
            ws.merge_cells("A1:C1")
            ws["A1"] = "AI QUALITY AUDIT REPORT"
            ws["A1"].font = Font(bold=True, size=14, color="4F46E5")
            ws["A1"].alignment = align_center
            
            ws["A3"] = "File Name:"
            ws["B3"] = record.filename
            ws["A4"] = "Date:"
            ws["B4"] = record.created_at.strftime("%Y-%m-%d %H:%M")
            ws["A5"] = "Overall Score:"
            ws["B5"] = f"{record.overall_score}/100"
            
            # 2. Scores Table
            ws["A7"] = "Metric"
            ws["B7"] = "Score (/10)"
            ws["C7"] = "Rationale"
            for cell in ["A7", "B7", "C7"]:
                ws[cell].font = header_font
                ws[cell].fill = header_fill
                ws[cell].alignment = align_center

            row = 8
            for metric, score in audit.get("scores", {}).items():
                ws.cell(row=row, column=1, value=metric.capitalize())
                ws.cell(row=row, column=2, value=score)
                ws.cell(row=row, column=3, value=audit.get("metric_justifications", {}).get(metric, ""))
                row += 1
                
            # 3. Key Findings
            row += 2
            ws.cell(row=row, column=1, value="Key Findings").font = header_font
            ws.cell(row=row, column=1).fill = header_fill
            row += 1
            for finding in audit.get("key_findings", []):
                ws.cell(row=row, column=1, value=finding)
                row += 1
                
            # 4. Transcript
            row += 2
            ws.cell(row=row, column=1, value="Transcript").font = header_font
            ws.cell(row=row, column=1).fill = header_fill
            row += 1
            for turn in transcript.get("dialogue", []):
                ws.cell(row=row, column=1, value=f"[{turn['speaker']}]: {turn['text']}")
                row += 1

            # Adjust column width
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 60

            # Stream buffer for response
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response['Content-Disposition'] = f'attachment; filename=audit_{audit_id}.xlsx'
            return response
            
        except AuditResult.DoesNotExist:
            return Response({"error": "Audit not found"}, status=404)
        except Exception as e:
            return Response({"error": str(e)}, status=500)
